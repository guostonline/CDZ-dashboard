from openpyxl import load_workbook
class SheetFix:
    def __init__(self,file_name) :
        self._day_work=24
        self.file_name=file_name
       
        pass
    
    
    @property
    def get_day_work(self):
       
        return self._day_work
    
    
    
    def fix_the_sheet(self):
        wb=load_workbook(self.file_name)
        sheet_ranges=wb["AGADIR"]
        wb.active
        day_work=sheet_ranges["C6"].value
        self._day_work=int(day_work.split("/",1)[0])
        print(wb.get_sheet_names() )
        for i in wb.get_sheet_names():
            if i != 'AGADIR':
                std=wb.get_sheet_by_name(i)
                wb.remove_sheet(std)
        try:
   
            sheet_ranges.unmerge_cells("A8:A9")
            sheet_ranges.unmerge_cells("B8:B9")
            sheet_ranges.unmerge_cells("D8:D9")
            sheet_ranges.unmerge_cells("F8:J8")
            sheet_ranges.unmerge_cells("K8:O8")
            sheet_ranges.delete_cols(1,2)
            sheet_ranges.delete_rows(1,8)
            sheet_ranges.delete_cols(7,6)
            sheet_ranges.delete_cols(7,2)
            sheet_ranges.delete_cols(3,1)
            sheet_ranges.delete_cols(7,1)
            
            
            sheet_ranges['A1']="Vendeur"
            sheet_ranges['B1']="Famille"
            sheet_ranges['E1']="Percent"
            
            
            
            
            for i in range(2,len(sheet_ranges["A"])+1): #problem ws.max
                sheet_ranges[f"C{i}"]=int(sheet_ranges[f"C{i}"].value)+int(sheet_ranges[f"F{i}"].value)
            for i in range(2,len(sheet_ranges["A"])+1):
                if sheet_ranges[f"D{i}"].value>0:
                    sheet_ranges[f"E{i}"]=(sheet_ranges[f"C{i}"].value/sheet_ranges[f"D{i}"].value)-1
                else : sheet_ranges[f"E{i}"]=0.0
            
        
            
          
            wb.save("finale.xlsx")
            
            print("Mod completed successfully")
    
        except Exception as e:print(e)